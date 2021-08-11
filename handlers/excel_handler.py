import os
from openpyxl import load_workbook
from utils.helpers import remove_directory_content

def load_KSK_list(KSK_file_path: str, wirelist_file_path: str):
    """Convert a KSK list excel sheet to an object based on a wire-list"""
    if not os.path.exists(KSK_file_path):
        raise ValueError("KSK path is invalid")
    
    KSK_wb = load_workbook(KSK_file_path)
    KSK_sheet = KSK_wb.active
    
    
    all_KSK = {}
    max_column = KSK_sheet.max_column - 1
    while(max_column > 0):
        data = []
        derivatives = []
        KSK_name = KSK_sheet[1][max_column].value
        for row in KSK_sheet:
            if (row[0].value):
                if row[max_column].value == 'X':
                     # remove derivative version then add it to the list
                     derivatives.append(row[0].value.split(' ')[0])

        wirelist_data = load_wire_list(wirelist_file_path)
        for derivative, connector, empty_cavity in wirelist_data:
            if derivative in derivatives:
                data.append((connector, empty_cavity))
        if data:
            all_KSK[KSK_name] = data
        max_column -= 1

    if not all_KSK:
        raise ValueError("Files fromat is invalid!")

    remove_directory_content('output')
    return all_KSK
    
    


def load_wire_list(file_path: str):
    """Load a wire-list excel sheet and return the neccessary data as a list"""
    if not os.path.exists(file_path):
        raise ValueError("Wire-list path is invalid")

    wirelist_wb = load_workbook(file_path)
    wirelist_sheet = wirelist_wb.active
    data = []  # data = [(derivative, connector, cavity), ...]
    
    for row in list(wirelist_sheet)[1:]:  # row by row
        # row[0] : Derivative
        # row[1] : Connector
        # row[2] : Cavity
        derivative = row[0].value.split(' ')[0]
        connector_id = row[1].value.strip()
        empty_cavity = row[2].value
        data.append((derivative, connector_id, empty_cavity))

    return data


